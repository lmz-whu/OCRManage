# -*- coding: utf-8 -*-

"""
Module implementing MainWindow.
"""
import os
import sqlite3

from PyQt4.QtCore import pyqtSlot
from PyQt4.QtGui import QMainWindow
from PyQt4 import QtGui
from PyQt4 import QtCore

from Ui_MainForm import Ui_MainWindow


class MainWindow(QMainWindow, Ui_MainWindow):
    """
    Class documentation goes here.
    """
    my_dbname = "OCRManage.db"


    def __init__(self, parent = None):
        """
        Constructor
        
        @param parent reference to the parent widget (QWidget)
        """
        super().__init__(parent)
        self.setupUi(self)

    @pyqtSlot()
    def on_pushButton_ShowRecordContent_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
#         raise NotImplementedError
        row = self.tableWidget_Records.currentRow()
        if row != -1:
            self.on_tableWidget_Records_cellDoubleClicked(row, 0)


    @pyqtSlot()
    def on_pushButton_LoadRecords_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
#         raise NotImplementedError
        self.tableWidget_Records.clear()
        # 设置表的抬头
        self.tableWidget_Records.setColumnCount(4)
        self.tableWidget_Records.setHorizontalHeaderLabels(["图片名称", "算法识别", "人工识别", "耗时/ms"])
        # 读取数据库中的数据，然后添加到表中
        my_conn = sqlite3.connect(self.my_dbname)
        my_cursor = my_conn.cursor()

        # 根据组合框中的选择项来确定当前应当加载哪些数据项
        cur_picType = self.comboBox_picTypes.currentText()
        surfix = str.format("and ocr_records.type='{0}'", cur_picType)
        if cur_picType == "所有项" or cur_picType == "":    # 如果什么都没选或者选择了"所有项"，则加载所有选项
            surfix = " "

        my_cursor.execute("SELECT ocr_records.pic_name,\
                                   ocr_records.ocr_text,\
                                   original_pictures.real_text,\
                                   ocr_records.consume_time\
                              FROM ocr_records,\
                                   original_pictures\
                             WHERE ocr_records.pic_name = original_pictures.pic_name %s" % surfix)
        for i, row_content in enumerate(my_cursor):    # 遍历每一行
            if self.tableWidget_Records.rowCount() <= i:    # 插入一行
                self.tableWidget_Records.insertRow(i)
            for j, col_content in enumerate(row_content):    # 遍历每一行的每一列
                newitem = QtGui.QTableWidgetItem(str(col_content))
                self.tableWidget_Records.setItem(i, j, newitem)

        my_conn.commit()
        my_conn.close()

    @pyqtSlot()
    def on_action_import_pic_into_db_triggered(self):
        """
        #将图片数据导入到数据库中，选择图片文件夹，用文件夹的名称表示图片的种类
        #然后图片文件夹下面所有哦的图片，然后保存在数据库中
        """
        # TODO: not implemented yet
#         raise NotImplementedError
        my_conn = sqlite3.connect(self.my_dbname)
        my_cursor = my_conn.cursor()
        my_dir = QtGui.QFileDialog.getExistingDirectory(self, '选择需要导入的图片文件夹', './')
        if my_dir == '':
            my_conn.close()
            return
        my_type = my_dir.split("\\")[-1]    # 设置图片的的种类,即文件夹的名称
        # 如果选择了文件夹路径，则开始处理
        succeedcount = 0    # 记录成功插入的数量
        failedcount = 0    # 记录插入失败的数量
        for picname in os.listdir(my_dir):
            # 判断文件是否是以图片后缀
            if str(picname).endswith((".jpg", ".png", ".bmp", ".gif", ".jpeg")):
                # 打开文件，以二进制方式打开，存储到数据库的blob字段
                with open(my_dir + "\\" + picname, "rb") as fp:
                    try:
                        # 插入数据库
                        my_cursor.execute("insert into picture values (?, ?, ?)", (my_type, picname, fp.read()))
                        succeedcount += 1
                    except Exception as e:
                        self.listWidget_log.addItem(" ".join([my_type, picname, "插入失败", str(e)]))
                        failedcount += 1
        self.listWidget_log.addItem(" ".join(["成功：", str(succeedcount), "失败", str(failedcount)]))

        my_conn.commit()
        my_conn.close()

    @pyqtSlot()
    def on_action_import_record_into_db_triggered(self):
        """
        Slot documentation goes here.
        """

        # TODO: not implemented yet
#         raise NotImplementedError
        txtname = QtGui.QFileDialog.getOpenFileName(self, '选择OCR检测结果文件', './', '*.txt')
        if txtname == '':
            return

        my_conn = sqlite3.connect(self.my_dbname)
        my_cursor = my_conn.cursor()

        succeedcount = 0    # 记录成功插入的数量
        failedcount = 0    # 记录插入失败的数量
        my_type = txtname.split("/")[-1].split(".")[0][:-1]    # 表示检测的图片的种类
        with  open(txtname, "r") as fp:
            while True:
                pic_name = ''
                D_detectret = []
                E_time = ''
                F_remark = ''

                # 首先读取一行，判断是否到文件底部
                line = str(fp.readline())
                if line == "":
                    break
#                 bool_write = False
                # 根据读取的第一行，判断是不是新的记录开始
                if line.startswith("*"):
                    pic_name = str(fp.readline()).strip()    # 读取图片的名称
                    # 读取中间的多余的两行
                    fp.readlines(2)
                    # 开始读取中间的识别结果，如果有识别结果，就开始循环读取结果，如果没有直接continue
                    line = str(fp.readline())
                    if line.startswith("识别结果"):
                        while True:
                            # 在读取到识别结果之后，开始读取--->所表示的识别结果，如果有结果，则添加到列表中，并设置可以写入xlsx
                            line = str(fp.readline())
                            if line.startswith("--->"):
#                                 bool_write = True
                                D_detectret.append(line[4:].strip())
                            else :
                                break
                    # 无论从何处退出，当前读取的行都是以“耗时”开始的一行
                    E_time = line.split("：")[-1][:-3]
                    # 插入数据，并在日志中添加插入的结果
                    try:
                        my_cursor.execute("insert into ocr_records values(?,?,?,?,?)",
                                          (my_type, pic_name, "; ".join(D_detectret), E_time, F_remark))
                        succeedcount += 1
                    except Exception as e:
                        self.listWidget_log.addItem(" ".join([" 插入失败", my_type, pic_name, "; ".join(D_detectret), E_time, F_remark, str(e)]))
                        failedcount += 1
        # 最后显示插入的结果，成功和失败的数量
        self.listWidget_log.addItem(" ".join(["成功：", str(succeedcount), "失败", str(failedcount)]))
        my_conn.commit()
        my_conn.close()


    @pyqtSlot()
    def on_action_quit_triggered(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
#         raise NotImplementedError
        self.close()


    @pyqtSlot(int, int)
    def on_tableWidget_Records_cellDoubleClicked(self, row, column):
        """
        双击单元格之后，会在停靠窗口中显示检测之后的图片和算法识别结果等信息
        显示的图片是算法处理之后的图片，来自数据库中的picture表格，
        由于处理后的图片名称和原始图片名称后缀不同，这里采用的前缀匹配
        """

        # TODO: not implemented yet
#         raise NotImplementedError

        pic_name_full = self.tableWidget_Records.item(row, 0).text()
        pic_name = pic_name_full.split(".")[0]
        ocr_text = self.tableWidget_Records.item(row, 1).text()
        real_text = self.tableWidget_Records.item(row, 2).text()
        consume_time = self.tableWidget_Records.item(row, 3).text()

        my_conn = sqlite3.connect(self.my_dbname)
        my_cursor = my_conn.cursor()
        try:
            my_cursor.execute("select image from picture where pic_name like ? ", (pic_name + "%",))
            pixmap = QtGui.QPixmap()
            pixmap.loadFromData(my_cursor.fetchone()[0])
            scene = QtGui.QGraphicsScene()
            scene.addPixmap(pixmap)
            self.graphicsView_OriginalPic.setScene(scene)
            self.label_TimeConsume.setText(consume_time + " ms")
            self.label_PictureName.setText(pic_name_full)
            self.textBrowser_OCRResult.clear()
            self.textBrowser_OCRResult.append(ocr_text)
            self.textBrowser_realtext.clear()
            self.textBrowser_realtext.append(real_text)
        except Exception as e:
            self.listWidget_log.addItem("显示记录错误：" + str(e))

        my_conn.close()

    @pyqtSlot()
    def on_action_import_real_text_triggered(self):
        """
        选择xlsx文件，该文件包含了，图片的文件名和人工识别的结果
        文件的第一个表单，必须命名为'Sheet1'
        文件格式：{A:图片名称1; B:图片名称2; C:目视识别结果; ...}
        """

        # TODO: not implemented yet
        from openpyxl import load_workbook

        xlsxname = QtGui.QFileDialog.getOpenFileName(self, '选择人工识别结果文件', './', '*.xlsx')
        if xlsxname == '':
            return

        wb = load_workbook(filename = xlsxname)
        ws = wb['Sheet1']

        succeedcount = 0    # 记录成功插入的数量
        failedcount = 0    # 记录插入失败的数量

        for i, row in enumerate(ws.rows):
            if i == 0 or row[0] == '':
                continue
            pic_name = "_".join([str(row[0].value).strip(), str(row[1].value).strip()])
#             print(row[2].value)
            real_text = str(row[2].value).strip() if row[2].value is not None else " "

            my_conn = sqlite3.connect(self.my_dbname)
            my_cursor = my_conn.cursor()

            try:
                my_cursor.execute("update original_pictures set real_text='%s' where pic_name='%s'" % (real_text, pic_name))
                if 0 == my_conn.total_changes:
                    failedcount += 1
                else:
                    succeedcount += 1
                    self.listWidget_log.addItem(" ".join(["成功：", pic_name, real_text]))
                my_conn.commit()
            except Exception as e:
                self.listWidget_log.addItem(" ".join(["错误：", pic_name, real_text, str(e)]))
                failedcount += 1

        self.listWidget_log.addItem(" ".join(["成功：", str(succeedcount), "失败", str(failedcount)]))
        my_conn.commit()
        my_conn.close()


    @pyqtSlot()
    def on_pushButton_Refresh_PicTypes_clicked(self):
        """
        Slot documentation goes here.
        """

        # TODO: not implemented yet
#         raise NotImplementedError
        self.comboBox_picTypes.clear()
        self.comboBox_picTypes.addItem("所有项")
        my_conn = sqlite3.connect(self.my_dbname)
        my_cursor = my_conn.cursor()

        my_cursor.execute("select type from ocr_records group by type")
        for row in my_cursor:
            self.comboBox_picTypes.addItem(str(row[0]))

        my_conn.close()



    @pyqtSlot()
    def on_pushButton_Choose_PicTypes_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
#         raise NotImplementedError
        self.on_pushButton_LoadRecords_clicked()

    @pyqtSlot()
    def on_action_import_original_pic_into_db_triggered(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        from openpyxl import load_workbook
        my_conn = sqlite3.connect(self.my_dbname)
        my_cursor = my_conn.cursor()
        my_dir = QtGui.QFileDialog.getExistingDirectory(self, '选择需要导入的图片文件夹', './')
        my_xlsxfile = QtGui.QFileDialog.getOpenFileName(self, '选择人工识别结果文件', './', '*.xlsx')
        if my_dir == '' or my_xlsxfile == '':
            my_conn.close()
            return

        wb = load_workbook(filename = my_xlsxfile)    # 加载xlsx文件
        ws = wb['Sheet1']

        my_type = my_dir.split("\\")[-1]    # 设置图片的的种类
        # 如果选择了文件夹路径，则开始处理
        succeedcount = 0    # 记录成功插入的数量
        failedcount = 0    # 记录插入失败的数量
        for picname in os.listdir(my_dir):
            # 判断文件是否是以图片后缀
            if str(picname).endswith((".jpg", ".png", ".bmp", ".gif", ".jpeg")):
                # 从xlsx中寻找相对应的图片文件，然后读取图片文件所对应的 目视识别结果
                # 由于不同xlsx文件中的文件名命名不同，这里方法也不同
                real_text = ""    # 表示目视识别的文字
                pic_prefix = int(picname.split("_")[0])    # 获取图片前缀，用来匹配图片
                for row in ws.rows:
                    if row[0] == '':
                        continue
                    try:
                        if int(str(row[0].value).strip()) == pic_prefix:    # 若果在xlsx文件中找到该图片名称，设置目视识别的文字
                            real_text = str(row[2].value).strip() if row[2].value is not None else " "
                            break
                    except Exception as e:
                        continue
                # 打开文件，以二进制方式打开，存储到数据库的blob字段
                with open(my_dir + "\\" + picname, "rb") as fp:
                    try:
                        # 插入数据库
                        my_cursor.execute("insert into original_pictures values (?, ?, ?, ?)", (my_type, picname, fp.read(), real_text))
                        succeedcount += 1
                    except Exception as e:
                        self.listWidget_log.addItem(" ".join([my_type, picname, "插入失败", str(e)]))
                        failedcount += 1
        self.listWidget_log.addItem(" ".join(["成功：", str(succeedcount), "失败", str(failedcount)]))

        my_conn.commit()
        my_conn.close()


    @pyqtSlot()
    def on_pushButton_update_ocrtext_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
#         raise NotImplementedError
        pic_name = self.label_PictureName.text().strip()
        ocr_text = self.textBrowser_OCRResult.toPlainText().strip()

        try:
            my_conn = sqlite3.connect(self.my_dbname)
            my_cursor = my_conn.cursor()
            my_cursor.execute("update ocr_records set ocr_text=? where pic_name=?", (ocr_text, pic_name))
            my_conn.commit()
            my_conn.close()
            self.listWidget_log.addItem("更新成功：" + pic_name + ocr_text)
        except Exception as e:
            self.listWidget_log.addItem("Error: " + str(e))

        newitem = QtGui.QTableWidgetItem(ocr_text)
        self.tableWidget_Records.setItem(self.tableWidget_Records.currentRow(), 1, newitem)


    @pyqtSlot()
    def on_pushButton_update_realtext_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
#         raise NotImplementedError
        pic_name = self.label_PictureName.text().strip()
        real_text = self.textBrowser_realtext.toPlainText().strip()

        try:
            my_conn = sqlite3.connect(self.my_dbname)
            my_cursor = my_conn.cursor()
            my_cursor.execute("update original_pictures set real_text=? where pic_name=?", (real_text, pic_name))
            my_conn.commit()
            my_conn.close()
            self.listWidget_log.addItem("更新成功：" + pic_name + real_text)
        except Exception as e:
            self.listWidget_log.addItem("Error: " + str(e))

        newitem = QtGui.QTableWidgetItem(real_text)
        self.tableWidget_Records.setItem(self.tableWidget_Records.currentRow(), 2, newitem)
